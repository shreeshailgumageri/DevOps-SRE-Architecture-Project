import boto3
import os
from botocore.exceptions import ClientError
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from collections import Counter
import argparse

# === Constants ===
DAYS_UNUSED = 90
CUTOFF_DATE = datetime.now(timezone.utc) - timedelta(days=DAYS_UNUSED)

# === Result Containers ===
unused_resources = []
risky_sgs = []
misconfigurations = []
resource_type_counter = Counter()
misconfig_type_counter = Counter()

# === Helper function to append and count ===
def add_unused_resource(profile, region, rtype, id_name):
    unused_resources.append({
        'Profile': profile,
        'Region': region,
        'ResourceType': rtype,
        'ID/Name': id_name
    })
    resource_type_counter[rtype] += 1

def add_misconfiguration(profile, region, resource_type, resource_id, issue, severity='HIGH'):
    misconfigurations.append({
        'Profile': profile,
        'Region': region,
        'ResourceType': resource_type,
        'ResourceId': resource_id,
        'Issue': issue,
        'Severity': severity
    })
    misconfig_type_counter[resource_type] += 1

# === Resource Check Functions ===
def check_resources(session, profile_name):
    profile_display = profile_name if profile_name else 'default'
    regions = session.get_available_regions('ec2')
    for region in regions:
        print(f"🔍 [{profile_display}] Scanning region: {region}")
        try:
            ec2 = session.client('ec2', region_name=region)
            lambda_client = session.client('lambda', region_name=region)
            cloudwatch = session.client('cloudwatch', region_name=region)
            rds = session.client('rds', region_name=region)

            # EC2 Instances
            reservations = ec2.describe_instances()['Reservations']
            for res in reservations:
                for inst in res['Instances']:
                    # Check for unused EC2 instances
                    if inst['State']['Name'] == 'stopped' and inst['LaunchTime'].replace(tzinfo=timezone.utc) < CUTOFF_DATE:
                        add_unused_resource(profile_name, region, 'EC2 Instance', inst['InstanceId'])
                    
                    # Check EC2 instance security misconfigurations
                    if inst['State']['Name'] != 'terminated':
                        # Check if instance is running without IMDSv2
                        metadata_options = inst.get('MetadataOptions', {})
                        if metadata_options.get('HttpTokens') != 'required':
                            add_misconfiguration(profile_name, region, 'EC2 Instance', inst['InstanceId'],
                                                "IMDSv2 (token-based metadata service) not required", 'MEDIUM')
                        
                        # Check for public IP
                        if 'PublicIpAddress' in inst:
                            add_misconfiguration(profile_name, region, 'EC2 Instance', inst['InstanceId'],
                                                f"Has public IP: {inst['PublicIpAddress']}", 'LOW')
                        
                        # Check if instance has detailed monitoring enabled
                        if inst.get('Monitoring', {}).get('State') != 'enabled':
                            add_misconfiguration(profile_name, region, 'EC2 Instance', inst['InstanceId'],
                                                "Detailed monitoring not enabled", 'LOW')

            # EBS Volumes
            volumes = ec2.describe_volumes()
            for vol in volumes['Volumes']:
                # Check for unused volumes
                if vol['State'] == 'available' and vol['CreateTime'].replace(tzinfo=timezone.utc) < CUTOFF_DATE:
                    add_unused_resource(profile_name, region, 'EBS Volume', vol['VolumeId'])
                
                # Check for unencrypted EBS volumes
                if not vol.get('Encrypted', False):
                    add_misconfiguration(profile_name, region, 'EBS Volume', vol['VolumeId'],
                                        "Volume is not encrypted", 'HIGH')
                
                # Check for volumes without recent snapshots
                try:
                    snapshots = ec2.describe_snapshots(Filters=[{'Name': 'volume-id', 'Values': [vol['VolumeId']]}])
                    has_recent_snapshot = False
                    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
                    
                    for snapshot in snapshots['Snapshots']:
                        if snapshot['StartTime'].replace(tzinfo=timezone.utc) > thirty_days_ago:
                            has_recent_snapshot = True
                            break
                    
                    if not has_recent_snapshot and vol['State'] == 'in-use':
                        add_misconfiguration(profile_name, region, 'EBS Volume', vol['VolumeId'],
                                            "No snapshot in last 30 days", 'MEDIUM')
                except Exception as e:
                    print(f"⚠️ Error checking snapshots for volume {vol['VolumeId']}: {e}")

            # Elastic IPs
            eips = ec2.describe_addresses()['Addresses']
            for eip in eips:
                if not any(key in eip for key in ['InstanceId', 'NetworkInterfaceId', 'AssociationId']):
                    add_unused_resource(profile_name, region, 'Elastic IP (Unused)', eip.get('PublicIp'))

            # Key Pairs
            key_pairs = ec2.describe_key_pairs()['KeyPairs']
            used_keys = {inst.get('KeyName') for res in reservations for inst in res['Instances'] if 'KeyName' in inst}
            for kp in key_pairs:
                if kp['KeyName'] not in used_keys:
                    add_unused_resource(profile_name, region, 'Key Pair (Unused)', kp['KeyName'])

            # Lambda
            get_unused_lambda_functions(lambda_client, cloudwatch, region, profile_name)

            # Security Groups
            find_risky_and_unused_sgs(session, ec2, region, profile_name)
            
            # RDS Databases
            check_rds_misconfigurations(rds, region, profile_name)

        except ClientError as e:
            print(f"❌ AWS error in {region}: {e}")
        except Exception as e:
            print(f"⚠️ Unexpected error in {region}: {e}")

    print(f"🔍 [{profile_name}] Scanning S3 Buckets globally")
    check_unused_s3_buckets(session, CUTOFF_DATE, profile_name)
    
    print(f"🔍 [{profile_name}] Scanning global IAM settings")
    check_iam_misconfigurations(session, profile_name)

def get_unused_lambda_functions(lambda_client, cloudwatch, region, profile_name):
    paginator = lambda_client.get_paginator('list_functions')
    for page in paginator.paginate():
        for function in page['Functions']:
            fn_name = function['FunctionName']
            metrics = cloudwatch.get_metric_statistics(
                Namespace='AWS/Lambda',
                MetricName='Invocations',
                Dimensions=[{'Name': 'FunctionName', 'Value': fn_name}],
                StartTime=CUTOFF_DATE,
                EndTime=datetime.now(timezone.utc),
                Period=86400,
                Statistics=['Sum']
            )
            if not metrics['Datapoints']:
                add_unused_resource(profile_name, region, 'Lambda (Unused)', fn_name)

def find_risky_and_unused_sgs(session, ec2, region, profile_name):
    all_sgs = ec2.describe_security_groups()['SecurityGroups']
    attached_sg_ids = set()

    # ENIs
    eni = ec2.describe_network_interfaces()['NetworkInterfaces']
    for iface in eni:
        for sg in iface.get('Groups', []):
            attached_sg_ids.add(sg['GroupId'])

    # Load Balancers (ALB/NLB)
    try:
        elbv2 = session.client('elbv2', region_name=region)
        lbs = elbv2.describe_load_balancers()['LoadBalancers']
        for lb in lbs:
            attached_sg_ids.update(lb.get('SecurityGroups', []))
    except Exception as e:
        print(f"⚠️ Error checking ALB/NLB SGs in {region}: {e}")

    # Classic Load Balancers
    try:
        elb = session.client('elb', region_name=region)
        clbs = elb.describe_load_balancers()['LoadBalancerDescriptions']
        for clb in clbs:
            attached_sg_ids.update(clb.get('SecurityGroups', []))
    except Exception as e:
        print(f"⚠️ Error checking CLB SGs in {region}: {e}")

    # RDS
    try:
        rds = session.client('rds', region_name=region)
        dbs = rds.describe_db_instances()['DBInstances']
        for db in dbs:
            for vpc_sg in db.get('VpcSecurityGroups', []):
                attached_sg_ids.add(vpc_sg['VpcSecurityGroupId'])
    except Exception as e:
        print(f"⚠️ Error checking RDS SGs in {region}: {e}")

    # ElastiCache
    try:
        elasticache = session.client('elasticache', region_name=region)
        clusters = elasticache.describe_cache_clusters(ShowCacheNodeInfo=True)['CacheClusters']
        for cluster in clusters:
            for sg in cluster.get('SecurityGroups', []):
                attached_sg_ids.add(sg['SecurityGroupId'])
    except Exception as e:
        print(f"⚠️ Error checking ElastiCache SGs in {region}: {e}")

    # Redshift
    try:
        redshift = session.client('redshift', region_name=region)
        clusters = redshift.describe_clusters()['Clusters']
        for cluster in clusters:
            for sg in cluster.get('VpcSecurityGroups', []):
                attached_sg_ids.add(sg['VpcSecurityGroupId'])
    except Exception as e:
        print(f"⚠️ Error checking Redshift SGs in {region}: {e}")

    # Lambda with VPC config
    try:
        lambda_client = session.client('lambda', region_name=region)
        paginator = lambda_client.get_paginator('list_functions')
        for page in paginator.paginate():
            for fn in page['Functions']:
                if 'VpcConfig' in fn:
                    attached_sg_ids.update(fn['VpcConfig'].get('SecurityGroupIds', []))
    except Exception as e:
        print(f"⚠️ Error checking Lambda SGs in {region}: {e}")

    for sg in all_sgs:
        sg_id = sg['GroupId']
        sg_name = sg.get('GroupName', '')
        if sg_id not in attached_sg_ids:
            add_unused_resource(profile_name, region, 'SecurityGroup (Unused)', f"{sg_id} ({sg_name})")

        for perm in sg.get('IpPermissions', []):
            for ip_range in perm.get('IpRanges', []):
                if ip_range.get('CidrIp') == '0.0.0.0/0':
                    port = perm.get('FromPort')
                    if port in [22, 3389]:
                        risky_sgs.append({
                            'Profile': profile_name,
                            'Region': region,
                            'SG ID': sg_id,
                            'Port': port,
                            'SG Name': sg_name
                        })
                    

def check_iam_misconfigurations(session, profile_name):
    try:
        iam = session.client('iam')
        
        # Check for account password policy
        try:
            password_policy = iam.get_account_password_policy()['PasswordPolicy']
            
            # Check minimum password length
            if password_policy.get('MinimumPasswordLength', 0) < 14:
                add_misconfiguration(profile_name, 'global', 'IAM', 'Password Policy',
                                    "Password minimum length less than 14 characters", 'MEDIUM')
                
            # Check password reuse prevention
            if not password_policy.get('PasswordReusePrevention', 0) or password_policy.get('PasswordReusePrevention') < 24:
                add_misconfiguration(profile_name, 'global', 'IAM', 'Password Policy',
                                    "Password reuse prevention not set to 24 passwords", 'MEDIUM')
                
            # Check password expiry
            if not password_policy.get('ExpirePasswords', False):
                add_misconfiguration(profile_name, 'global', 'IAM', 'Password Policy',
                                    "Password expiration not enabled", 'MEDIUM')
                
            # Check for complexity requirements
            if not password_policy.get('RequireSymbols', False):
                add_misconfiguration(profile_name, 'global', 'IAM', 'Password Policy',
                                    "Password policy doesn't require symbols", 'MEDIUM')
                
            if not password_policy.get('RequireNumbers', False):
                add_misconfiguration(profile_name, 'global', 'IAM', 'Password Policy',
                                    "Password policy doesn't require numbers", 'MEDIUM')
                
            if not password_policy.get('RequireUppercaseCharacters', False):
                add_misconfiguration(profile_name, 'global', 'IAM', 'Password Policy',
                                    "Password policy doesn't require uppercase characters", 'MEDIUM')
                
            if not password_policy.get('RequireLowercaseCharacters', False):
                add_misconfiguration(profile_name, 'global', 'IAM', 'Password Policy',
                                    "Password policy doesn't require lowercase characters", 'MEDIUM')
                
        except ClientError as e:
            if 'NoSuchEntity' in str(e):
                add_misconfiguration(profile_name, 'global', 'IAM', 'Password Policy',
                                   "No account password policy set", 'HIGH')
            else:
                print(f"\u274c Error checking password policy: {e}")
        
        # Check IAM users
        users = iam.list_users()['Users']
        for user in users:
            username = user['UserName']
            
            # Check for console access without MFA
            try:
                login_profile = iam.get_login_profile(UserName=username)
                # User has console access, check for MFA
                mfa_devices = iam.list_mfa_devices(UserName=username)['MFADevices']
                if not mfa_devices:
                    add_misconfiguration(profile_name, 'global', 'IAM User', username,
                                      "User has console access without MFA", 'HIGH')
            except ClientError as e:
                if 'NoSuchEntity' not in str(e):
                    print(f"\u26a0\ufe0f Error checking login profile for {username}: {e}")
            
            # Check for access keys
            access_keys = iam.list_access_keys(UserName=username)['AccessKeyMetadata']
            for key in access_keys:
                key_id = key['AccessKeyId']
                if key['Status'] == 'Active':
                    # Check key age
                    key_created = key['CreateDate']
                    key_age = (datetime.now(timezone.utc) - key_created.replace(tzinfo=timezone.utc)).days
                    
                    if key_age > 90:
                        add_misconfiguration(profile_name, 'global', 'IAM User', f"{username} (Key: {key_id})",
                                          f"Access key is {key_age} days old (>90 days)", 'HIGH')
        
        # Check for IAM policies directly attached to users
        for user in users:
            username = user['UserName']
            attached_policies = iam.list_attached_user_policies(UserName=username)['AttachedPolicies']
            if attached_policies:
                policy_names = [p['PolicyName'] for p in attached_policies]
                add_misconfiguration(profile_name, 'global', 'IAM User', username,
                                  f"Has directly attached policies: {', '.join(policy_names)}", 'MEDIUM')
        
        # Check for unused roles
        roles = iam.list_roles()['Roles']
        for role in roles:
            role_name = role['RoleName']
            # Skip service-linked roles
            if role_name.startswith('AWSServiceRole') or 'service-role' in role.get('Path', ''):
                continue
                
            # Get role last used
            try:
                role_last_used = iam.get_role(RoleName=role_name)['Role'].get('RoleLastUsed', {})
                if 'LastUsedDate' not in role_last_used:
                    add_misconfiguration(profile_name, 'global', 'IAM Role', role_name,
                                      "Role has never been used", 'LOW')
                else:
                    last_used = role_last_used['LastUsedDate'].replace(tzinfo=timezone.utc)
                    days_since_used = (datetime.now(timezone.utc) - last_used).days
                    if days_since_used > 90:
                        add_misconfiguration(profile_name, 'global', 'IAM Role', role_name,
                                          f"Role not used for {days_since_used} days", 'LOW')
            except ClientError as e:
                print(f"\u26a0\ufe0f Error checking role usage for {role_name}: {e}")
        
    except ClientError as e:
        print(f"\u274c Error checking IAM configurations: {e}")
    except Exception as e:
        print(f"\u26a0\ufe0f Unexpected error checking IAM configurations: {e}")


def check_rds_misconfigurations(rds_client, region, profile_name):
    try:
        # Get all RDS instances
        instances = rds_client.describe_db_instances()
        
        for db in instances['DBInstances']:
            db_id = db['DBInstanceIdentifier']
            
            # Check for encryption
            if not db.get('StorageEncrypted', False):
                add_misconfiguration(profile_name, region, 'RDS Instance', db_id,
                                   "Database storage not encrypted", 'HIGH')
            
            # Check for public accessibility
            if db.get('PubliclyAccessible', False):
                add_misconfiguration(profile_name, region, 'RDS Instance', db_id,
                                   "Database is publicly accessible", 'HIGH')
                
            # Check for automated backups
            if db.get('BackupRetentionPeriod', 0) == 0:
                add_misconfiguration(profile_name, region, 'RDS Instance', db_id,
                                   "Automated backups not enabled", 'MEDIUM')
            
            # Check for multi-AZ (production databases)
            if not db.get('MultiAZ', False):
                add_misconfiguration(profile_name, region, 'RDS Instance', db_id,
                                   "Multi-AZ not enabled", 'LOW')
                
            # Check if enhanced monitoring is enabled
            if 'MonitoringInterval' not in db or db['MonitoringInterval'] == 0:
                add_misconfiguration(profile_name, region, 'RDS Instance', db_id,
                                   "Enhanced monitoring not enabled", 'LOW')
                
            # Check for automatic minor version upgrades
            if not db.get('AutoMinorVersionUpgrade', True):
                add_misconfiguration(profile_name, region, 'RDS Instance', db_id,
                                   "Automatic minor version upgrades disabled", 'LOW')
                
            # Check deletion protection
            if not db.get('DeletionProtection', False):
                add_misconfiguration(profile_name, region, 'RDS Instance', db_id,
                                   "Deletion protection not enabled", 'MEDIUM')
            
    except ClientError as e:
        print(f"❌ Error checking RDS instances in {region}: {e}")
    except Exception as e:
        print(f"⚠️ Unexpected error checking RDS instances in {region}: {e}")


def check_unused_s3_buckets(session, cutoff_date, profile_name):
    s3 = session.client('s3')
    buckets = s3.list_buckets().get('Buckets', [])
    for bucket in buckets:
        bucket_name = bucket['Name']
        try:
            loc = s3.get_bucket_location(Bucket=bucket_name).get('LocationConstraint')
            bucket_region = loc if loc else 'us-east-1'
            regional_s3 = session.client('s3', region_name=bucket_region)
            objects = regional_s3.list_objects_v2(Bucket=bucket_name, MaxKeys=1000)

            if 'Contents' not in objects:
                add_unused_resource(profile_name, bucket_region, 'S3 Bucket (Unused)', bucket_name)
                continue

            all_old = all(obj['LastModified'].replace(tzinfo=timezone.utc) < cutoff_date for obj in objects['Contents'])
            if all_old:
                add_unused_resource(profile_name, bucket_region, 'S3 Bucket (Unused)', bucket_name)

            # Check for S3 bucket misconfigurations
            check_s3_bucket_security(regional_s3, bucket_name, bucket_region, profile_name)

        except ClientError as e:
            print(f"❌ Error checking bucket {bucket_name}: {e}")

def check_s3_bucket_security(s3_client, bucket_name, region, profile_name):
    # Check for public access settings
    try:
        # Check bucket ACL for public access
        acl = s3_client.get_bucket_acl(Bucket=bucket_name)
        for grant in acl.get('Grants', []):
            grantee = grant.get('Grantee', {})
            if grantee.get('URI') == 'http://acs.amazonaws.com/groups/global/AllUsers' or \
               grantee.get('URI') == 'http://acs.amazonaws.com/groups/global/AuthenticatedUsers':
                add_misconfiguration(profile_name, region, 'S3 Bucket', bucket_name, 
                                     f"Public {grant.get('Permission')} permission granted", 'HIGH')

        # Check bucket policy
        try:
            policy = s3_client.get_bucket_policy(Bucket=bucket_name)
            if '"Principal":"*"' in str(policy) or '"Principal":{"AWS":"*"}' in str(policy):
                add_misconfiguration(profile_name, region, 'S3 Bucket', bucket_name, 
                                     "Bucket policy contains public access", 'HIGH')
        except ClientError as e:
            # No policy is actually secure
            if not 'NoSuchBucketPolicy' in str(e):
                print(f"Error checking bucket policy for {bucket_name}: {e}")
                
        # Check encryption
        try:
            encryption = s3_client.get_bucket_encryption(Bucket=bucket_name)
            rules = encryption.get('ServerSideEncryptionConfiguration', {}).get('Rules', [])
            if not rules:
                add_misconfiguration(profile_name, region, 'S3 Bucket', bucket_name, 
                                     "No default encryption", 'MEDIUM')
        except ClientError as e:
            if 'ServerSideEncryptionConfigurationNotFoundError' in str(e):
                add_misconfiguration(profile_name, region, 'S3 Bucket', bucket_name, 
                                     "No default encryption", 'MEDIUM')
            else:
                print(f"Error checking bucket encryption for {bucket_name}: {e}")
                
        # Check block public access settings
        try:
            public_access_block = s3_client.get_public_access_block(Bucket=bucket_name)
            config = public_access_block.get('PublicAccessBlockConfiguration', {})
            for setting, enabled in config.items():
                if not enabled:
                    add_misconfiguration(profile_name, region, 'S3 Bucket', bucket_name, 
                                         f"Public access block setting '{setting}' disabled", 'HIGH')
        except ClientError as e:
            if 'NoSuchPublicAccessBlockConfiguration' in str(e):
                add_misconfiguration(profile_name, region, 'S3 Bucket', bucket_name, 
                                     "No public access block configuration", 'HIGH')
            else:
                print(f"Error checking bucket public access block for {bucket_name}: {e}")
                
        # Check versioning (optional but recommended)
        try:
            versioning = s3_client.get_bucket_versioning(Bucket=bucket_name)
            if versioning.get('Status') != 'Enabled':
                add_misconfiguration(profile_name, region, 'S3 Bucket', bucket_name, 
                                     "Versioning not enabled", 'LOW')
        except ClientError as e:
            print(f"Error checking bucket versioning for {bucket_name}: {e}")

    except ClientError as e:
        print(f"Error checking security settings for bucket {bucket_name}: {e}")


def export_to_excel():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "UnusedResources"
    ws1.append(['Profile', 'Region', 'ResourceType', 'ID/Name'])
    for r in unused_resources:
        ws1.append([r['Profile'], r['Region'], r['ResourceType'], r['ID/Name']])

    ws2 = wb.create_sheet(title="RiskySecurityGroups")
    ws2.append(['Profile', 'Region', 'SG ID', 'Port', 'SG Name'])
    for r in risky_sgs:
        ws2.append([r['Profile'], r['Region'], r['SG ID'], r['Port'], r['SG Name']])
        
    ws3 = wb.create_sheet(title="SecurityMisconfigurations")
    ws3.append(['Profile', 'Region', 'Resource Type', 'Resource ID', 'Issue', 'Severity'])
    for r in misconfigurations:
        ws3.append([r['Profile'], r['Region'], r['ResourceType'], r['ResourceId'], r['Issue'], r['Severity']])

    filename = f"aws_security_assessment_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Exported results to {filename}")

def print_summary():
    print("\n📊 === Summary Report ===")
    print(f"🔸 Total Unused Resources: {len(unused_resources)}")
    for rtype, count in resource_type_counter.items():
        print(f"   • {rtype}: {count}")
    print(f"\n🔸 Total Risky Security Groups: {len(risky_sgs)}")
    
    # Group misconfigurations by severity
    high_severity = [m for m in misconfigurations if m['Severity'] == 'HIGH']
    medium_severity = [m for m in misconfigurations if m['Severity'] == 'MEDIUM']
    low_severity = [m for m in misconfigurations if m['Severity'] == 'LOW']
    
    print(f"\n🔸 Total Security Misconfigurations: {len(misconfigurations)}")
    print(f"   • HIGH severity: {len(high_severity)}")
    print(f"   • MEDIUM severity: {len(medium_severity)}")
    print(f"   • LOW severity: {len(low_severity)}")
    
    # Show top resource types with misconfigurations
    print("\n   Top misconfigured resources:")
    for rtype, count in misconfig_type_counter.most_common(5):
        print(f"   • {rtype}: {count}")
    
    # If there are high severity findings, show some of them
    if high_severity:
        print("\n   Sample HIGH severity findings:")
        for finding in high_severity[:5]:  # Show top 5 high severity findings
            print(f"   • {finding['ResourceType']} {finding['ResourceId']}: {finding['Issue']}")
            
    print("\n🚨 Review the full report in the Excel file for all findings.")


# === Main Execution ===
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scan AWS accounts for unused and risky resources")
    parser.add_argument('--profiles', nargs='+', required=False, help="Optional: List of AWS CLI profile names to scan")
    args = parser.parse_args()

    if args.profiles:
        profiles = args.profiles
    else:
        # Use default session when no profiles are provided
        profiles = [None]

    for profile in profiles:
        try:
            # Ensure AWS SDK loads config file
            os.environ['AWS_SDK_LOAD_CONFIG'] = '1'
            
            # Config file is mounted at /home/go/.aws/config in the pod
            config_path = '/home/go/.aws/config'
            
            if os.path.exists(config_path):
                print(f"🚀 Using AWS config file at: {config_path}")
                os.environ['AWS_CONFIG_FILE'] = config_path
            else:
                print(f"⚠️ AWS config file not found at {config_path}")
            
            if profile:
                print(f"🚀 Logging in using profile: {profile}")
                session = boto3.Session(profile_name=profile)
            else:
                # Try to get the default profile from the config file
                print("🚀 Using default profile from config file or environment credentials")
                session = boto3.Session()
            check_resources(session, profile)
        except Exception as e:
            print(f"❌ Failed for profile {profile}: {e}")

    print_summary()
    export_to_excel()